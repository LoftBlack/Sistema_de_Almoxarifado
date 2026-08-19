import io

from flask import Blueprint, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from models import db
from models.item import Item
from routes.auth import login_required

items_bp = Blueprint("items", __name__)


# ---------------------------------------------------------------------------
# Página principal (protegida por login)
# ---------------------------------------------------------------------------
@items_bp.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html")


# ---------------------------------------------------------------------------
# API: listar itens
# ---------------------------------------------------------------------------
@items_bp.route("/api/itens", methods=["GET"])
@login_required
def listar_itens():
    itens = Item.query.order_by(Item.nome.asc()).all()
    return jsonify([item.to_dict() for item in itens])


# ---------------------------------------------------------------------------
# API: criar item
# ---------------------------------------------------------------------------
@items_bp.route("/api/itens", methods=["POST"])
@login_required
def criar_item():
    dados = request.get_json(silent=True) or {}

    nome = (dados.get("nome") or "").strip()
    localizacao = (dados.get("localizacao") or "").strip()
    finalidade = (dados.get("finalidade") or "").strip()
    observacoes = (dados.get("observacoes") or "").strip()
    quantidade = dados.get("quantidade")

    erros = validar_item(nome, localizacao, finalidade, quantidade)
    if erros:
        return jsonify({"erros": erros}), 400

    item = Item(
        nome=nome,
        quantidade=int(quantidade),
        localizacao=localizacao,
        finalidade=finalidade,
        observacoes=observacoes,
    )
    db.session.add(item)
    db.session.commit()

    return jsonify(item.to_dict()), 201


# ---------------------------------------------------------------------------
# API: atualizar item
# ---------------------------------------------------------------------------
@items_bp.route("/api/itens/<int:item_id>", methods=["PUT"])
@login_required
def atualizar_item(item_id):
    item = Item.query.get_or_404(item_id)
    dados = request.get_json(silent=True) or {}

    nome = (dados.get("nome") or "").strip()
    localizacao = (dados.get("localizacao") or "").strip()
    finalidade = (dados.get("finalidade") or "").strip()
    quantidade = dados.get("quantidade")

    erros = validar_item(nome, localizacao, finalidade, quantidade)
    if erros:
        return jsonify({"erros": erros}), 400

    item.nome = nome
    item.quantidade = int(quantidade)
    item.localizacao = localizacao
    item.finalidade = finalidade
    item.observacoes = observacoes
    db.session.commit()

    return jsonify(item.to_dict())


# ---------------------------------------------------------------------------
# API: excluir item
# ---------------------------------------------------------------------------
@items_bp.route("/api/itens/<int:item_id>", methods=["DELETE"])
@login_required
def excluir_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({"sucesso": True})


# ---------------------------------------------------------------------------
# API: exportar itens para Excel (.xlsx)
# ---------------------------------------------------------------------------
@items_bp.route("/api/itens/exportar", methods=["GET"])
@login_required
def exportar_itens():
    itens = Item.query.order_by(Item.nome.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Almoxarifado"

    cabecalhos = ["Nome do Item", "Quantidade", "Localização", "Finalidade/Categoria"]
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color="6B4226", end_color="6B4226", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(cabecalhos) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for item in itens:
        ws.append([item.nome, item.quantidade, item.localizacao, item.finalidade])

    # Ajusta a largura das colunas automaticamente.
    for col_cells in ws.columns:
        valores = [str(c.value) for c in col_cells if c.value is not None]
        largura = (max(len(v) for v in valores) + 4) if valores else 15
        ws.column_dimensions[col_cells[0].column_letter].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="almoxarifado.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Validação básica compartilhada entre criar/atualizar
# ---------------------------------------------------------------------------
def validar_item(nome, localizacao, finalidade, quantidade):
    """Retorna uma lista de mensagens de erro. Lista vazia significa que passou."""
    erros = []

    if not nome:
        erros.append("O nome do item é obrigatório.")
    if not localizacao:
        erros.append("A localização é obrigatória.")
    if not finalidade:
        erros.append("A finalidade/categoria é obrigatória.")

    if quantidade is None or str(quantidade).strip() == "":
        erros.append("A quantidade é obrigatória.")
    else:
        try:
            quantidade_int = int(quantidade)
            if quantidade_int < 0:
                erros.append("A quantidade não pode ser negativa.")
        except (ValueError, TypeError):
            erros.append("A quantidade deve ser um número inteiro.")

    return erros
